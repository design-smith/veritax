from __future__ import annotations

import csv
import io
import re
import uuid
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path

from sqlalchemy.ext.asyncio import AsyncSession

from .extraction_schemas import fact_type_rule, schema_entry
from .extraction_store import (
    EntityMentionInput,
    ExpectedFieldInput,
    ExtractedFactInput,
    FactSourceInput,
    RunInput,
    add_expected_field,
    add_entity_mention,
    add_extracted_fact,
    extraction_fingerprint,
    get_or_create_extraction_run,
)
from .models import Document, Source
from .source_locators import validate_source_quote

RUNNER_VERSION = "financial-table-extractor-v1"
SCHEMA_KEY = "financial_table"

HEADER_ALIASES = {
    "account_code": {"accountcode", "accountno", "accountnumber", "glaccount", "account"},
    "account_name": {"accountname", "accountdescription", "description", "name"},
    "debit": {"debit", "dr"},
    "credit": {"credit", "cr"},
    "amount": {"amount", "balance", "netbalance", "value"},
    "date": {"date", "postingdate", "transactiondate", "invoicedate"},
    "invoice_number": {"invoice", "invoicenumber", "invoiceid", "documentnumber"},
    "period": {"fiscalyear", "period", "fy", "year"},
    "entity_name": {"entity", "legalentity", "company"},
    "counterparty_name": {"counterparty", "customer", "supplier", "relatedparty"},
}


@dataclass(frozen=True)
class TableRow:
    cells: list[str]
    locator: str
    quote: str


async def extract_trial_balance_document(
    session: AsyncSession,
    *,
    document: Document,
    data: bytes,
    classification_type: str,
    classification_version: str,
) -> object:
    return await _extract_financial_table_document(
        session,
        document=document,
        data=data,
        classification_type=classification_type,
        classification_version=classification_version,
        table_kind="trial_balance",
    )


async def extract_general_ledger_document(
    session: AsyncSession,
    *,
    document: Document,
    data: bytes,
    classification_type: str,
    classification_version: str,
) -> object:
    return await _extract_financial_table_document(
        session,
        document=document,
        data=data,
        classification_type=classification_type,
        classification_version=classification_version,
        table_kind="general_ledger",
    )


async def extract_invoice_population_document(
    session: AsyncSession,
    *,
    document: Document,
    data: bytes,
    classification_type: str,
    classification_version: str,
) -> object:
    return await _extract_financial_table_document(
        session,
        document=document,
        data=data,
        classification_type=classification_type,
        classification_version=classification_version,
        table_kind="invoice_population",
    )


async def _extract_financial_table_document(
    session: AsyncSession,
    *,
    document: Document,
    data: bytes,
    classification_type: str,
    classification_version: str,
    table_kind: str,
) -> object:
    schema = schema_entry(SCHEMA_KEY)
    schema_version = str(schema["schema_version"])
    rows = _read_rows(document.original_filename, data)
    header_index, columns = _find_header(rows, table_kind)
    missing_columns = _missing_columns(columns, table_kind)
    row_facts = [] if missing_columns else _facts_from_rows(rows[header_index + 1:], columns, table_kind)
    status = "extracted" if row_facts and not missing_columns else "partially_extracted"
    fingerprint = extraction_fingerprint(
        document_hash=document.content_hash,
        classification_type=classification_type,
        classification_version=classification_version,
        schema_version=schema_version,
        runner_version=RUNNER_VERSION,
        model_version="deterministic",
    )
    result = await get_or_create_extraction_run(
        session,
        RunInput(
            engagement_id=await _engagement_id(session, document),
            document_id=document.id,
            schema_key=SCHEMA_KEY,
            schema_version=schema_version,
            classification_type=classification_type,
            classification_version=classification_version,
            runner_version=RUNNER_VERSION,
            model_version="deterministic",
            fingerprint=fingerprint,
            status=status,
            diagnostics={
                "missing_columns": missing_columns,
                "rows_seen": max(0, len(rows) - header_index - 1),
                "rows_extracted": len({fact["row"].locator for fact in row_facts}),
            },
        ),
    )
    if result.reused:
        return result.run

    for missing in missing_columns:
        await add_expected_field(
            session,
            result.run.id,
            ExpectedFieldInput(field_name=missing, status="missing", reason="Could not map a financial table column."),
        )
    for item in row_facts:
        row = item["row"]
        region = validate_source_quote(
            document.original_filename,
            document.content_type,
            data,
            row.quote,
            locator=row.locator,
        )
        await _add_fact(
            session,
            result.run.id,
            document.id,
            schema_version,
            fact_type=item["fact_type"],
            value_raw=item["value_raw"],
            value_normalized=item.get("value_normalized"),
            period=item.get("period"),
            scope_level=item["scope_level"],
            source=FactSourceInput(
                document_id=document.id,
                page=region.page,
                locator=region.locator,
                quote=row.quote,
            ),
        )
    return result.run


def _read_rows(filename: str, data: bytes) -> list[TableRow]:
    ext = Path(filename).suffix.lower()
    if ext in {".xlsx", ".xls"}:
        return _xlsx_rows(data)
    return _csv_rows(data)


def _csv_rows(data: bytes) -> list[TableRow]:
    text = data.decode("utf-8", errors="ignore")
    rows = []
    for index, row in enumerate(csv.reader(io.StringIO(text)), start=1):
        cells = [str(cell).strip() for cell in row]
        if any(cells):
            rows.append(TableRow(cells=cells, locator=f"row {index}", quote="\t".join(cells)))
    return rows


def _xlsx_rows(data: bytes) -> list[TableRow]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[TableRow] = []
    for ws in wb.worksheets:
        for index, row in enumerate(ws.iter_rows(values_only=True), start=1):
            cells = [str(cell).strip() for cell in row if cell is not None]
            if cells:
                rows.append(TableRow(cells=cells, locator=f"{ws.title}!R{index}", quote="\t".join(cells)))
    return rows


def _find_header(rows: list[TableRow], table_kind: str) -> tuple[int, dict[str, int]]:
    best_index = 0
    best_map: dict[str, int] = {}
    for index, row in enumerate(rows):
        mapped = _map_headers(row.cells)
        if len(mapped) > len(best_map):
            best_index = index
            best_map = mapped
        if not _missing_columns(mapped, table_kind):
            return index, mapped
    return best_index, best_map


def _map_headers(cells: list[str]) -> dict[str, int]:
    mapped: dict[str, int] = {}
    for index, cell in enumerate(cells):
        normal = _normal_header(cell)
        for key, aliases in HEADER_ALIASES.items():
            if normal in aliases and key not in mapped:
                mapped[key] = index
    return mapped


def _missing_columns(columns: dict[str, int], table_kind: str) -> list[str]:
    missing = []
    if table_kind in {"trial_balance", "general_ledger"} and "account_code" not in columns and "account_name" not in columns:
        missing.append("account")
    if table_kind == "general_ledger" and "date" not in columns:
        missing.append("date")
    if table_kind == "invoice_population" and "invoice_number" not in columns:
        missing.append("invoice")
    if table_kind == "invoice_population" and "date" not in columns:
        missing.append("date")
    if "amount" not in columns and "debit" not in columns and "credit" not in columns:
        missing.append("amount")
    return missing


def _facts_from_rows(rows: list[TableRow], columns: dict[str, int], table_kind: str) -> list[dict]:
    out: list[dict] = []
    for row in rows:
        if table_kind == "invoice_population":
            out.extend(_invoice_row_facts(row, columns))
            continue
        out.extend(_accounting_row_facts(row, columns, table_kind))
    return out


def _accounting_row_facts(row: TableRow, columns: dict[str, int], table_kind: str) -> list[dict]:
    out: list[dict] = []
    account_code = _cell(row, columns, "account_code")
    account_name = _cell(row, columns, "account_name")
    amount_raw = _amount_raw(row, columns)
    amount = _parse_decimal(amount_raw)
    if amount is None:
        return out
    account_value = " ".join(v for v in (account_code, account_name) if v)
    period = _cell(row, columns, "period") or None
    out.append({
        "row": row,
        "fact_type": "gl_account",
        "value_raw": account_value,
        "value_normalized": _normalize_text(account_value),
        "period": period,
        "scope_level": "local_entity",
    })
    amount_type = _amount_fact_type(account_value) if table_kind == "trial_balance" else "transaction_amount"
    out.append({
        "row": row,
        "fact_type": amount_type,
        "value_raw": amount_raw,
        "value_normalized": _format_decimal(amount),
        "period": period,
        "scope_level": "local_entity" if table_kind == "trial_balance" else "transaction",
    })
    date = _cell(row, columns, "date")
    if date:
        out.append({
            "row": row,
            "fact_type": "transaction_date",
            "value_raw": date,
            "value_normalized": date,
            "period": period,
            "scope_level": "transaction",
        })
    out.extend(_party_facts(row, columns, period))
    return out


def _invoice_row_facts(row: TableRow, columns: dict[str, int]) -> list[dict]:
    out: list[dict] = []
    amount_raw = _amount_raw(row, columns)
    amount = _parse_decimal(amount_raw)
    invoice = _cell(row, columns, "invoice_number")
    date = _cell(row, columns, "date")
    if amount is None:
        return out
    out.append({
        "row": row,
        "fact_type": "invoice_number",
        "value_raw": invoice,
        "value_normalized": _normalize_text(invoice),
        "period": None,
        "scope_level": "transaction",
    })
    out.append({
        "row": row,
        "fact_type": "transaction_date",
        "value_raw": date,
        "value_normalized": date,
        "period": None,
        "scope_level": "transaction",
    })
    out.append({
        "row": row,
        "fact_type": "invoice_amount",
        "value_raw": amount_raw,
        "value_normalized": _format_decimal(amount),
        "period": None,
        "scope_level": "transaction",
    })
    out.extend(_party_facts(row, columns, None))
    return out


def _party_facts(row: TableRow, columns: dict[str, int], period: str | None) -> list[dict]:
    out: list[dict] = []
    entity = _cell(row, columns, "entity_name")
    if entity:
        out.append({
            "row": row,
            "fact_type": "entity_name",
            "value_raw": entity,
            "value_normalized": _normalize_text(entity),
            "period": period,
            "scope_level": "local_entity",
        })
    counterparty = _cell(row, columns, "counterparty_name")
    if counterparty:
        out.append({
            "row": row,
            "fact_type": "counterparty_name",
            "value_raw": counterparty,
            "value_normalized": _normalize_text(counterparty),
            "period": period,
            "scope_level": "counterparty",
        })
    return out


def _cell(row: TableRow, columns: dict[str, int], key: str) -> str:
    index = columns.get(key)
    if index is None or index >= len(row.cells):
        return ""
    return row.cells[index].strip()


def _amount_raw(row: TableRow, columns: dict[str, int]) -> str:
    if "amount" in columns:
        return _cell(row, columns, "amount")
    debit = _cell(row, columns, "debit")
    credit = _cell(row, columns, "credit")
    credit_num = _parse_decimal(credit)
    if credit_num is not None and credit_num != 0:
        return credit
    return debit


def _parse_decimal(value: str) -> Decimal | None:
    cleaned = str(value or "").strip()
    if not cleaned:
        return None
    negative = cleaned.startswith("(") and cleaned.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", cleaned)
    if not cleaned:
        return None
    try:
        amount = Decimal(cleaned)
    except InvalidOperation:
        return None
    return -amount if negative else amount


def _format_decimal(value: Decimal) -> str:
    formatted = format(value.normalize(), "f")
    return "0" if formatted == "-0" else formatted


def _amount_fact_type(account: str) -> str:
    normal = account.casefold()
    if any(word in normal for word in ("revenue", "sales", "income")):
        return "revenue"
    if any(word in normal for word in ("expense", "cost", "fee")):
        return "expense"
    return "transaction_amount"


def _normal_header(value: str) -> str:
    return re.sub(r"[^a-z0-9]", "", value.casefold())


def _normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", value.casefold()).strip()


async def _add_fact(
    session: AsyncSession,
    run_id: uuid.UUID,
    document_id: uuid.UUID,
    schema_version: str,
    *,
    fact_type: str,
    value_raw: str,
    value_normalized: str | None,
    period: str | None,
    scope_level: str,
    source: FactSourceInput,
) -> None:
    rule = fact_type_rule(SCHEMA_KEY, fact_type)
    fact = await add_extracted_fact(
        session,
        run_id,
        document_id,
        ExtractedFactInput(
            schema_key=SCHEMA_KEY,
            schema_version=schema_version,
            fact_type=fact_type,
            value_raw=value_raw,
            value_normalized=value_normalized,
            value_type=str(rule["value_type"]),
            period=period,
            scope_level=scope_level,
            resolution_status="unresolved" if rule["value_type"] == "entity_ref" else "not_required",
        ),
        sources=[source],
    )
    if rule["value_type"] == "entity_ref":
        await add_entity_mention(
            session,
            run_id,
            document_id,
            fact.id,
            EntityMentionInput(
                raw_name=value_raw,
                role=fact_type,
                locator=source.locator,
                quote=source.quote,
            ),
        )


async def _engagement_id(session: AsyncSession, document: Document) -> uuid.UUID:
    source = await session.get(Source, document.source_id)
    if source is None:
        raise ValueError(f"source not found for document {document.id}")
    return source.engagement_id
