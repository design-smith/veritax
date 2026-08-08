from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

from .classification_store import ClassificationInput, scope_fingerprint
from .config import settings
from .evidence_taxonomy import document_type_entry, load_taxonomy
from .processing import extract_text
from .requirements import available_jurisdictions

CLASSIFIER_VERSION = "rules-v1"
PREVIEW_CHARS = 4_000
YEAR_RE = re.compile(r"\b(?:FY\s*)?(20\d{2})\b", re.IGNORECASE)
LEGAL_ENTITY_SUFFIXES = (" b.v.", " bv", " gmbh", " limited", " ltd", " llc", " inc", " corp", " plc")
FALLBACK_SYSTEM_PROMPT = (
    "You classify transfer-pricing source documents from short previews. "
    "Return only JSON with primary_type, observed_signals, supporting_quotes, candidate_tags, "
    "and candidate_scope_values. primary_type must be one of the allowed_document_types. "
    "Do not return confidence."
)


def _parse_fallback_json(text: str) -> dict:
    import json_repair

    data = json_repair.loads(text)
    if not isinstance(data, dict):
        raise RuntimeError("classification fallback returned non-object JSON")
    return data


class AnthropicClassificationFallback:
    def __init__(self) -> None:
        self._client = None
        self._model = settings.assessment_model

    def _get_client(self):
        if self._client is None:
            import anthropic

            self._client = anthropic.Anthropic(api_key=settings.anthropic_api_key)
        return self._client

    def classify(self, payload: dict) -> dict:
        resp = self._get_client().messages.create(
            model=self._model,
            max_tokens=700,
            system=FALLBACK_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": json.dumps(payload, ensure_ascii=True)}],
        )
        text = "".join(getattr(block, "text", "") for block in resp.content)
        return _parse_fallback_json(text)


class DeepSeekClassificationFallback:
    def __init__(self) -> None:
        self._client = None
        self._model = settings.deepseek_model

    def _get_client(self):
        if self._client is None:
            from openai import OpenAI

            self._client = OpenAI(api_key=settings.deepseek_api_key, base_url=settings.deepseek_base_url)
        return self._client

    def classify(self, payload: dict) -> dict:
        resp = self._get_client().chat.completions.create(
            model=self._model,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": FALLBACK_SYSTEM_PROMPT},
                {"role": "user", "content": json.dumps(payload, ensure_ascii=True)},
            ],
            max_tokens=700,
        )
        return _parse_fallback_json(resp.choices[0].message.content or "{}")


def _decode(data: bytes) -> str:
    return data.decode("utf-8", errors="ignore")


def _csv_preview(data: bytes) -> str:
    text = _decode(data)
    sample = "\n".join(text.splitlines()[:12])
    try:
        rows = list(csv.reader(io.StringIO(sample)))
    except csv.Error:
        return sample[:PREVIEW_CHARS]
    return "\n".join(" | ".join(cell.strip() for cell in row[:20]) for row in rows)[:PREVIEW_CHARS]


def _xlsx_preview(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    parts: list[str] = []
    for ws in wb.worksheets[:5]:
        parts.append(f"sheet: {ws.title} rows: {ws.max_row or '?'} cols: {ws.max_column or '?'}")
        for row in ws.iter_rows(max_row=12, values_only=True):
            cells = [str(cell).strip() for cell in row[:20] if cell is not None and str(cell).strip()]
            if cells:
                parts.append(" | ".join(cells))
    return "\n".join(parts)[:PREVIEW_CHARS]


def _docx_preview(data: bytes) -> str:
    from docx import Document as Docx

    doc = Docx(io.BytesIO(data))
    parts: list[str] = []
    for paragraph in doc.paragraphs[:24]:
        text = paragraph.text.strip()
        if text:
            parts.append(text)
        if sum(len(part) for part in parts) >= PREVIEW_CHARS:
            break
    return "\n".join(parts)[:PREVIEW_CHARS]


def _preview(filename: str, content_type: str | None, data: bytes) -> str:
    ext = Path(filename).suffix.lower()
    ctype = (content_type or "").lower()
    if ext == ".csv" or "csv" in ctype:
        return _csv_preview(data)
    if ext in (".xlsx", ".xls") or "spreadsheet" in ctype or "excel" in ctype:
        return _xlsx_preview(data)
    if ext == ".docx" or "wordprocessingml" in ctype:
        return _docx_preview(data)
    return extract_text(filename, content_type, data)[:PREVIEW_CHARS]


def _normal(value: str | None) -> str:
    return (value or "").strip().casefold()


def _find_fiscal_year(text: str) -> str | None:
    match = YEAR_RE.search(text)
    return f"FY{match.group(1)}" if match else None


def _call_fallback(llm_fallback: Any, payload: dict) -> tuple[dict | None, list[str]]:
    errors: list[str] = []
    for _ in range(2):
        try:
            result = llm_fallback.classify(payload)
            primary_type = str(result.get("primary_type") or "")
            document_type_entry(primary_type)
            return result, errors
        except Exception as exc:  # noqa: BLE001 - retry once, then degrade to Unknown
            errors.append(str(exc)[:500])
    return None, errors


def _score_types(text: str) -> tuple[str, int, list[str]]:
    taxonomy = load_taxonomy()
    best_type = "Unknown"
    best_score = 0
    best_signals: list[str] = []
    for entry in taxonomy["document_types"]:
        if entry["document_type"] == "Unknown":
            continue
        signals = [s for s in entry.get("deterministic_signals", []) if _normal(s) in text]
        score = len(signals) * 20
        if _normal(entry["document_type"]) in text:
            score += 25
            signals.append(entry["document_type"])
        if score > best_score:
            best_type = entry["document_type"]
            best_score = score
            best_signals = list(dict.fromkeys(signals))
    return best_type, best_score, best_signals


def _has_other_entity_signal(text: str, entity_name: str | None) -> bool:
    if not entity_name or _normal(entity_name) in text:
        return False
    return "entity" in text and any(suffix in text for suffix in LEGAL_ENTITY_SUFFIXES)


def classify_document_bytes(
    *,
    filename: str,
    content_type: str | None,
    content_hash: str,
    data: bytes,
    entity_name: str | None,
    jurisdictions: list[str],
    fiscal_year: str | None,
    llm_fallback: Any = None,
) -> ClassificationInput:
    preview = _preview(filename, content_type, data)
    blob = _normal(f"{filename}\n{preview}")
    document_type, score, signals = _score_types(blob)
    fallback_errors: list[str] = []
    llm_quotes: list[str] = []
    scope_values: dict = {}

    if llm_fallback is not None and (document_type == "Unknown" or score < 40):
        fallback_result, fallback_errors = _call_fallback(
            llm_fallback,
            {
                "filename": filename,
                "content_type": content_type,
                "preview": preview,
                "deterministic_type": document_type,
                "deterministic_score": score,
                "deterministic_signals": signals,
                "entity_name": entity_name,
                "jurisdictions": jurisdictions,
                "fiscal_year": fiscal_year,
                "allowed_document_types": [entry["document_type"] for entry in load_taxonomy()["document_types"]],
            },
        )
        if fallback_result is not None:
            document_type = str(fallback_result.get("primary_type"))
            observed = [str(v) for v in fallback_result.get("observed_signals", []) if str(v).strip()]
            signals = list(dict.fromkeys([*signals, *observed]))
            score = max(score, min(95, 40 + len(signals) * 10))
            llm_quotes = [str(v) for v in fallback_result.get("supporting_quotes", []) if str(v).strip()]
            scope_values = fallback_result.get("candidate_scope_values") or {}
        elif score < 40:
            document_type = "Unknown"
            score = 0
    entry = document_type_entry(document_type)

    found_entity = entity_name if entity_name and _normal(entity_name) in blob else None
    jurisdiction_blob = blob.replace(_normal(entity_name), "") if entity_name else blob
    known_jurisdiction = next((j for j in available_jurisdictions() if _normal(j) in jurisdiction_blob), None)
    found_jurisdiction = (
        known_jurisdiction
        if known_jurisdiction and any(_normal(known_jurisdiction) == _normal(j) for j in jurisdictions)
        else None
    )
    found_year = _find_fiscal_year(blob)
    if not found_entity and entity_name and _normal(scope_values.get("entity")) == _normal(entity_name):
        found_entity = entity_name
    other_entity = _has_other_entity_signal(blob, entity_name)
    if not found_entity and scope_values.get("entity") and entity_name:
        other_entity = _normal(scope_values.get("entity")) != _normal(entity_name)
    if not found_jurisdiction:
        found_jurisdiction = next(
            (j for j in jurisdictions if _normal(scope_values.get("jurisdiction")) == _normal(j)),
            None,
        )
        if not found_jurisdiction and scope_values.get("jurisdiction"):
            known_jurisdiction = str(scope_values.get("jurisdiction"))
    if not found_year and scope_values.get("fiscal_year"):
        found_year = _find_fiscal_year(str(scope_values.get("fiscal_year")))
    expected_year = fiscal_year.strip().upper() if fiscal_year else None

    validation = {
        "entity": "pass" if found_entity else ("fail" if other_entity else "unknown"),
        "jurisdiction": (
            "pass" if found_jurisdiction
            else "fail" if known_jurisdiction and jurisdictions
            else "unknown"
        ),
        "fiscal_year": (
            "pass" if found_year and expected_year and found_year == expected_year
            else "fail" if found_year and expected_year and found_year != expected_year
            else "unknown"
        ),
    }

    if "fail" in validation.values():
        relevance = "out_of_scope"
    elif document_type == "Unknown":
        relevance = "unknown"
    elif all(validation[k] == "pass" for k in ("entity", "jurisdiction", "fiscal_year")):
        relevance = "relevant"
    else:
        relevance = "partially_relevant"

    tags = list(entry.get("allowed_tags", []))
    if found_year:
        tags.append(found_year)
    return ClassificationInput(
        document_type=document_type,
        classification_score=score,
        classification_state="accepted" if document_type != "Unknown" and score >= 40 else "unknown",
        relevance=relevance,
        tags=list(dict.fromkeys(tags)),
        entity=found_entity,
        jurisdiction=found_jurisdiction,
        fiscal_year=found_year,
        language="English" if re.search(r"\b(the|and|agreement|account)\b", blob) else None,
        document_status="Executed" if "executed" in blob else None,
        version=None,
        source_validation_result=validation,
        deterministic_signals=signals,
        llm_supporting_quotes=llm_quotes,
        candidate_requirements=list(entry.get("candidate_requirement_categories", [])),
        candidate_extractors=list(entry.get("candidate_extractors", [])),
        scope_fingerprint=scope_fingerprint(
            document_hash=content_hash,
            entity_name=entity_name,
            jurisdictions=jurisdictions,
            fiscal_year=fiscal_year,
            classifier_version=CLASSIFIER_VERSION,
        ),
        classifier_version=CLASSIFIER_VERSION,
        diagnostics={
            "preview_chars": len(preview),
            "filename": filename,
            "fallback_used": bool(llm_quotes),
            "fallback_errors": fallback_errors,
        },
        taxonomy_version=str(load_taxonomy()["taxonomy_version"]),
    )


def unknown_classification(
    *,
    filename: str,
    content_hash: str,
    entity_name: str | None,
    jurisdictions: list[str],
    fiscal_year: str | None,
    error: str,
) -> ClassificationInput:
    return ClassificationInput(
        document_type="Unknown",
        classification_score=0,
        classification_state="unknown",
        relevance="unknown",
        tags=[],
        source_validation_result={"entity": "unknown", "jurisdiction": "unknown", "fiscal_year": "unknown"},
        deterministic_signals=[],
        llm_supporting_quotes=[],
        candidate_requirements=[],
        candidate_extractors=[],
        scope_fingerprint=scope_fingerprint(
            document_hash=content_hash,
            entity_name=entity_name,
            jurisdictions=jurisdictions,
            fiscal_year=fiscal_year,
            classifier_version=CLASSIFIER_VERSION,
        ),
        classifier_version=CLASSIFIER_VERSION,
        diagnostics={"filename": filename, "error": error[:500]},
        taxonomy_version=str(load_taxonomy()["taxonomy_version"]),
    )
