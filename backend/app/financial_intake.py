"""Financial data intake & normalization (Class 3 §7-11).

Parse XLSX/CSV into canonical rows with immutable raw-cell capture + deterministic provenance. Default column
detection only (deterministic exact/normalized header match) — real column mapping incl. saved mappings + LLM
suggestions is S3 (§12-14). Never fabricates a value: an unparseable amount stays `None` (S4 diagnoses it, §15).
"""
from __future__ import annotations

import csv
import hashlib
import io
from dataclasses import dataclass

from openpyxl import load_workbook

# Canonical field -> recognised source headers (normalized). S3 extends with aliases + saved client mappings.
_DEFAULT_MAP: dict[str, set[str]] = {
    "account_code": {"account_code", "gl code", "gl_code", "account code", "account", "code", "acct", "acct code"},
    "account_name": {"account_name", "gl description", "account name", "description", "name", "acct name", "account description"},
    "amount": {"amount", "balance", "value", "fy actual", "actual", "amount (local)", "local amount", "amount local"},
    "currency": {"currency", "currency code", "ccy", "curr"},
    "cost_center": {"cost_center", "cost center", "cc", "costcentre", "cost centre"},
    "business_unit": {"business_unit", "business unit", "bu", "segment"},
    "counterparty": {"counterparty", "ic partner", "intercompany partner", "partner", "counterparty entity"},
    "period": {"period", "fiscal_year", "fiscal year", "fy", "year"},
}

CANONICAL_FIELDS = tuple(_DEFAULT_MAP)


def _norm(h: object) -> str:
    return " ".join(str(h if h is not None else "").strip().lower().split())


def detect_columns(headers: list[str]) -> dict[str, str]:
    """{canonical_field: source_header} for headers we recognise (default detection; S3 refines/overrides)."""
    out: dict[str, str] = {}
    for src in headers:
        n = _norm(src)
        if not n:
            continue
        for field_name, aliases in _DEFAULT_MAP.items():
            if field_name not in out and n in aliases:
                out[field_name] = src
                break
    # Fallback for the critical `amount` column — TBs/GLs label it many ways (e.g. "FY24 Actual", "Closing
    # Balance"). Deterministic substring match; real per-client mapping is S3.
    if "amount" not in out:
        for src in headers:
            if any(k in _norm(src) for k in ("actual", "amount", "balance")):
                out["amount"] = src
                break
    return out


def parse_amount(v: object) -> float | None:
    """Parse a cell to a number, tolerating thousands separators, currency symbols, and accounting negatives.
    Returns None when it can't be parsed — never a fabricated 0 (§2, §15)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = s.startswith("(") and s.endswith(")")           # accounting negatives: (1,234)
    s = s.strip("()").replace(",", "").replace("$", "").strip()
    try:
        n = float(s)
    except ValueError:
        return None
    return -n if neg else n


def derive_from_mapping(raw: dict, mapping: dict[str, str]) -> dict:
    """Canonical field values from the immutable raw cells, given {canonical_field: source_header}.

    This is what makes a remap cheap (§9): rows are re-derived from `raw` when the mapping changes — no re-upload.
    """
    def cell(field: str) -> str | None:
        header = mapping.get(field)
        if not header:
            return None
        v = raw.get(header)
        s = None if v is None else str(v).strip()
        return s or None

    amount_header = mapping.get("amount")
    return {
        "account_code": cell("account_code"),
        "account_name": cell("account_name"),
        "amount": parse_amount(raw.get(amount_header)) if amount_header else None,
        "currency": cell("currency"),
        "cost_center": cell("cost_center"),
        "business_unit": cell("business_unit"),
        "counterparty": cell("counterparty"),
        "period": cell("period"),
    }


def header_signature(headers: list[str]) -> str:
    """Stable signature of a header set (order-independent, normalized) — the key for reusing a saved mapping."""
    norm = "|".join(sorted(h for h in (_norm(x) for x in headers) if h))
    return hashlib.sha256(norm.encode("utf-8")).hexdigest()[:32]


@dataclass
class ParsedRow:
    row_index: int          # 1-based source data-row number (excludes the header row)
    raw: dict               # header -> original cell (string) — immutable capture (§9)
    account_code: str | None
    account_name: str | None
    amount: float | None
    currency: str | None
    cost_center: str | None
    business_unit: str | None
    counterparty: str | None
    period: str | None


@dataclass
class ParsedFinancials:
    sheet: str | None
    headers: list[str]
    detected: dict[str, str]   # canonical_field -> source header
    rows: list[ParsedRow]


def _build_rows(sheet: str | None, headers: list[str], records: list[list[object]]) -> ParsedFinancials:
    detected = detect_columns(headers)
    idx = {f: headers.index(src) for f, src in detected.items()}

    def cell(row: list[object], f: str) -> object:
        i = idx.get(f)
        return row[i] if i is not None and i < len(row) else None

    def text(row: list[object], f: str) -> str | None:
        v = cell(row, f)
        s = None if v is None else str(v).strip()
        return s or None

    rows: list[ParsedRow] = []
    n = 0
    for record in records:
        if all(c is None or str(c).strip() == "" for c in record):
            continue    # skip fully-blank rows (never a fabricated row)
        n += 1
        raw = {headers[i]: ("" if i >= len(record) or record[i] is None else str(record[i]))
               for i in range(len(headers))}
        rows.append(ParsedRow(
            row_index=n, raw=raw,
            account_code=text(record, "account_code"),
            account_name=text(record, "account_name"),
            amount=parse_amount(cell(record, "amount")),
            currency=text(record, "currency"),
            cost_center=text(record, "cost_center"),
            business_unit=text(record, "business_unit"),
            counterparty=text(record, "counterparty"),
            period=text(record, "period"),
        ))
    return ParsedFinancials(sheet=sheet, headers=headers, detected=detected, rows=rows)


def _parse_csv(data: bytes) -> ParsedFinancials:
    text = data.decode("utf-8-sig", errors="replace")
    reader = list(csv.reader(io.StringIO(text)))
    rows = [r for r in reader if any((c or "").strip() for c in r)]
    if not rows:
        return ParsedFinancials(sheet=None, headers=[], detected={}, rows=[])
    headers = [str(h).strip() for h in rows[0]]
    return _build_rows(None, headers, [list(r) for r in rows[1:]])


def _parse_xlsx(data: bytes) -> ParsedFinancials:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    grid = [list(r) for r in ws.iter_rows(values_only=True)]
    sheet = ws.title
    wb.close()
    grid = [r for r in grid if any(c is not None and str(c).strip() for c in r)]
    if not grid:
        return ParsedFinancials(sheet=sheet, headers=[], detected={}, rows=[])
    headers = [("" if h is None else str(h).strip()) for h in grid[0]]
    return _build_rows(sheet, headers, grid[1:])


def parse_financial_file(filename: str, data: bytes) -> ParsedFinancials:
    """Parse an uploaded financial file into canonical rows. v1 accepts XLSX/XLS(M)/CSV (§7)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        return _parse_csv(data)
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        return _parse_xlsx(data)
    raise ValueError(f"unsupported financial file type: {filename!r} (v1 accepts XLSX/CSV)")
