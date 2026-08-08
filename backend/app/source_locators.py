from __future__ import annotations

import csv
import io
import re
from dataclasses import dataclass
from pathlib import Path


@dataclass(frozen=True)
class SourceRegion:
    locator: str
    text: str
    page: int | None = None


def source_regions(filename: str, content_type: str | None, data: bytes) -> list[SourceRegion]:
    ext = Path(filename).suffix.lower().lstrip(".")
    if ext == "pdf":
        return _pdf_regions(data)
    if ext == "docx":
        return _docx_regions(data)
    if ext in {"xlsx", "xls"}:
        return _xlsx_regions(data)
    if ext == "csv":
        return _csv_regions(data)
    return _text_regions(data)


def validate_source_quote(
    filename: str,
    content_type: str | None,
    data: bytes,
    quote: str,
    *,
    page: int | None = None,
    locator: str | None = None,
) -> SourceRegion:
    if not quote or not quote.strip():
        raise ValueError("quote is required")
    regions = source_regions(filename, content_type, data)
    ext = Path(filename).suffix.lower().lstrip(".")
    if ext == "pdf" and page is None and any(region.text.strip() for region in regions):
        raise ValueError("page is required for PDF source quotes")
    candidates = regions
    if page is not None:
        candidates = [region for region in candidates if region.page == page]
        if not candidates:
            raise ValueError(f"page {page} not found")
    if locator is not None:
        candidates = [region for region in candidates if region.locator == locator]
        if not candidates:
            raise ValueError(f"locator {locator!r} not found")
    for region in candidates:
        if _contains_quote(region.text, quote):
            return region
    raise ValueError("quote not found in source")


def _contains_quote(text: str, quote: str) -> bool:
    return _squash(quote) in _squash(text)


def _squash(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def _pdf_regions(data: bytes) -> list[SourceRegion]:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return [
        SourceRegion(locator=f"page {index}", page=index, text=page.extract_text() or "")
        for index, page in enumerate(reader.pages, start=1)
    ]


def _docx_regions(data: bytes) -> list[SourceRegion]:
    from docx import Document

    doc = Document(io.BytesIO(data))
    return [
        SourceRegion(locator=f"paragraph {index}", text=paragraph.text)
        for index, paragraph in enumerate(doc.paragraphs, start=1)
    ]


def _xlsx_regions(data: bytes) -> list[SourceRegion]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    regions: list[SourceRegion] = []
    for sheet in wb.worksheets:
        for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            cells = [str(cell) for cell in row if cell is not None]
            if cells:
                regions.append(SourceRegion(locator=f"{sheet.title}!R{index}", text="\t".join(cells)))
    return regions


def _csv_regions(data: bytes) -> list[SourceRegion]:
    text = data.decode("utf-8", errors="ignore")
    return [
        SourceRegion(locator=f"row {index}", text="\t".join(row))
        for index, row in enumerate(csv.reader(io.StringIO(text)), start=1)
        if any(cell.strip() for cell in row)
    ]


def _text_regions(data: bytes) -> list[SourceRegion]:
    text = data.decode("utf-8", errors="ignore")
    return [
        SourceRegion(locator=f"line {index}", text=line)
        for index, line in enumerate(text.splitlines(), start=1)
    ]
