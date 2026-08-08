from __future__ import annotations

import csv
import io
import logging
import re
import shutil
import subprocess
import tempfile
from collections.abc import Iterator
from pathlib import Path

from .config import settings

log = logging.getLogger("veritax")

# ── Text extraction by file type ─────────────────────────────────────────────
_VTT_SRT_TS = re.compile(r"^\d{1,2}:\d{2}:\d{2}[.,]\d{3}\s*-->.*$")
_SRT_INDEX = re.compile(r"^\d+$")


def _ext(filename: str) -> str:
    return filename.rsplit(".", 1)[-1].lower() if "." in filename else ""


def _substantive_chars(text: str) -> int:
    return len(re.sub(r"\s+", "", text or ""))


def _has_enough_text(text: str) -> bool:
    return _substantive_chars(text) >= settings.ocr_min_text_chars


def extract_text(filename: str, content_type: str | None, data: bytes) -> str:
    ext = _ext(filename)
    try:
        if ext == "pdf":
            return _pdf(data)
        if ext == "docx":
            return _docx(data)
        if ext in ("xlsx", "xls"):
            return _xlsx(data)
        if ext == "csv":
            return _csv(data)
        if ext in ("vtt", "srt"):
            return _cues(data)
        if ext in ("txt", "md", "text", ""):
            return data.decode("utf-8", errors="ignore")
    except Exception:
        # Extraction is best-effort for findability; a parse failure yields no chunks, not a 500.
        return ""
    # Unknown binary types → no text (0 chunks).
    return ""


def _pdf(data: bytes) -> str:
    text = _pdf_text_layer(data)
    if _has_enough_text(text):
        return text
    if not settings.ocr_enabled:
        log.info(
            "pdf extract: text layer too thin (%d chars), OCR disabled",
            _substantive_chars(text),
        )
        return text
    try:
        log.info(
            "pdf extract: text layer too thin (%d chars), trying OCR",
            _substantive_chars(text),
        )
        ocr_text = _ocr_pdf(data)
    except Exception as exc:  # noqa: BLE001 - keep indexing failure actionable, not a 500
        log.warning("pdf extract: OCR unavailable or failed: %s", exc)
        return text
    if _substantive_chars(ocr_text) > _substantive_chars(text):
        log.info("pdf extract: OCR produced %d substantive chars", _substantive_chars(ocr_text))
        return ocr_text
    log.info("pdf extract: OCR did not improve extraction")
    return text


def _pdf_text_layer(data: bytes) -> str:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(data))
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _ocr_pdf(data: bytes) -> str:
    command = settings.ocr_command.strip()
    if not command:
        raise RuntimeError("OCR command is not configured")
    if shutil.which(command) is None:
        raise RuntimeError(f"OCR command '{command}' is not installed")

    with tempfile.TemporaryDirectory(prefix="veritax-ocr-") as workdir:
        base = Path(workdir)
        input_path = base / "input.pdf"
        output_path = base / "output.pdf"
        sidecar_path = base / "sidecar.txt"
        input_path.write_bytes(data)
        cmd = [
            command,
            "--force-ocr",
            "--jobs",
            "1",
            "--output-type",
            "pdf",
            "--sidecar",
            str(sidecar_path),
        ]
        if settings.ocr_language:
            cmd.extend(["--language", settings.ocr_language])
        cmd.extend([str(input_path), str(output_path)])
        result = subprocess.run(
            cmd,
            cwd=workdir,
            capture_output=True,
            text=True,
            timeout=settings.ocr_timeout_seconds,
            check=False,
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(f"OCR command failed with exit {result.returncode}: {detail[:500]}")
        if not sidecar_path.exists():
            raise RuntimeError("OCR command did not produce sidecar text")
        return sidecar_path.read_text(encoding="utf-8", errors="ignore")


def _docx(data: bytes) -> str:
    from docx import Document as Docx

    doc = Docx(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs)


def _xlsx(data: bytes) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append("\t".join(cells))
    return "\n".join(rows)


def _csv(data: bytes) -> str:
    text = data.decode("utf-8", errors="ignore")
    return "\n".join("\t".join(r) for r in csv.reader(io.StringIO(text)))


def _cues(data: bytes) -> str:
    """Strip WEBVTT/SRT timestamps and indices, keep spoken text."""
    lines: list[str] = []
    for raw in data.decode("utf-8", errors="ignore").splitlines():
        line = raw.strip()
        if not line or line == "WEBVTT" or _VTT_SRT_TS.match(line) or _SRT_INDEX.match(line):
            continue
        lines.append(line)
    return "\n".join(lines)


# ── Chunking ─────────────────────────────────────────────────────────────────
# ponytail: naive fixed-size word window with overlap. Good enough for findability search now;
# swap for structural/semantic chunking when retrieval quality matters.
def iter_chunks(text: str, words_per_chunk: int = 600, overlap: int = 80) -> Iterator[str]:
    if words_per_chunk <= 0:
        return
    overlap = max(0, min(overlap, words_per_chunk - 1))
    window: list[str] = []
    new_since_yield = 0
    for match in re.finditer(r"\S+", text):
        window.append(match.group(0))
        new_since_yield += 1
        if len(window) >= words_per_chunk:
            yield " ".join(window)
            window = window[-overlap:] if overlap else []
            new_since_yield = 0
    if window and new_since_yield > 0:
        yield " ".join(window)


def chunk(text: str, words_per_chunk: int = 600, overlap: int = 80) -> list[str]:
    return list(iter_chunks(text, words_per_chunk, overlap))
